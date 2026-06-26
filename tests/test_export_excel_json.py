"""
Validation tests: RepositoryOverview and UpgradePath data in Excel/JSON exports.
Covers:
  - Excel export produces a valid .xlsx with Repository Overview + Upgrade Path sheets
  - Excel sheets contain KPI data and per-job upgrade findings
  - JSON export produces valid JSON with structured repositoryOverview + upgradePath keys
  - build_report_pack produces excel_path / json_path alongside docx/html/pdf
"""
import json
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.tiap.documentation.report_pack_generator import (
    build_report_pack,
    build_report_pack_sections,
)
from app.reports.excel_export import write_complete_assessment_excel
from app.reports.json_export import write_complete_assessment_json


def _make_job(name: str, source_version: str = "Talend 6", components=None):
    return {
        "job_data": {
            "job_name": name,
            "source_version": source_version,
            "components": [{"component_type": c} for c in (components or ["tMap"])],
        }
    }


SAMPLE_JOBS = [
    _make_job("Job_Alpha", "Talend 6", ["tMap", "tFTPGet", "tJavaFlex"]),
    _make_job("Job_Beta",  "Talend 7", ["tMap", "tLogRow"]),
]


# ── Excel export ──────────────────────────────────────────────────────────────

class TestExcelExport:

    def _build_excel(self) -> str:
        sections = build_report_pack_sections(SAMPLE_JOBS)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        write_complete_assessment_excel(path, sections, SAMPLE_JOBS)
        return path

    def test_excel_file_created(self):
        path = self._build_excel()
        try:
            assert os.path.exists(path)
            assert os.path.getsize(path) > 100
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_excel_is_valid_workbook(self):
        path = self._build_excel()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert len(wb.sheetnames) > 0
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_excel_contains_repository_overview_sheet(self):
        path = self._build_excel()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "Repository Overview" in wb.sheetnames
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_excel_contains_upgrade_path_sheet(self):
        path = self._build_excel()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "Upgrade Path" in wb.sheetnames
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_excel_repository_overview_has_kpi_rows(self):
        path = self._build_excel()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb["Repository Overview"]
            values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
            assert any("Total Jobs" in v for v in values)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_excel_upgrade_path_has_job_names(self):
        path = self._build_excel()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb["Upgrade Path"]
            values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
            assert any("Job_Alpha" in v for v in values)
            assert any("Job_Beta" in v for v in values)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_excel_other_sections_present_as_sheets(self):
        path = self._build_excel()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            assert "Executive Summary" in wb.sheetnames
        finally:
            if os.path.exists(path):
                os.unlink(path)


# ── JSON export ───────────────────────────────────────────────────────────────

class TestJSONExport:

    def _build_json(self) -> dict:
        sections = build_report_pack_sections(SAMPLE_JOBS)
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = f.name
        try:
            write_complete_assessment_json(path, sections, SAMPLE_JOBS)
            with open(path, encoding="utf-8") as fh:
                return json.load(fh)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_json_is_valid_and_parseable(self):
        data = self._build_json()
        assert isinstance(data, dict)

    def test_json_contains_repository_overview_key(self):
        data = self._build_json()
        assert "repositoryOverview" in data

    def test_json_contains_upgrade_path_key(self):
        data = self._build_json()
        assert "upgradePath" in data

    def test_json_repository_overview_total_jobs(self):
        data = self._build_json()
        assert data["repositoryOverview"].get("totalJobs") == 2

    def test_json_upgrade_path_has_jobs_list(self):
        data = self._build_json()
        assert "jobs" in data["upgradePath"]
        job_names = {j["jobName"] for j in data["upgradePath"]["jobs"]}
        assert "Job_Alpha" in job_names
        assert "Job_Beta" in job_names

    def test_json_upgrade_path_source_and_target_version(self):
        data = self._build_json()
        assert data["upgradePath"]["sourceVersion"] == "Talend 6"
        assert data["upgradePath"]["targetVersion"] == "Talend 8"

    def test_json_sections_key_present(self):
        data = self._build_json()
        assert "sections" in data
        assert "Repository Overview" in data["sections"]
        assert "Upgrade Path" in data["sections"]

    def test_json_unknown_source_version_graceful(self):
        sections = build_report_pack_sections([])
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = f.name
        try:
            write_complete_assessment_json(path, sections, [])
            with open(path, encoding="utf-8") as fh:
                data = json.load(fh)
            assert data["upgradePath"]["available"] is False
        finally:
            if os.path.exists(path):
                os.unlink(path)


# ── build_report_pack integration ──────────────────────────────────────────────

class TestBuildReportPackAllFormats:

    def test_build_report_pack_creates_all_export_paths(self):
        with tempfile.TemporaryDirectory() as tmp:
            result = build_report_pack(all_jobs=SAMPLE_JOBS, output_dir=tmp)
            for key in ("docx_path", "html_path", "pdf_path", "excel_path", "json_path"):
                assert key in result
                assert os.path.exists(result[key])

    def test_build_report_pack_excel_nontrivial_size(self):
        with tempfile.TemporaryDirectory() as tmp:
            result = build_report_pack(all_jobs=SAMPLE_JOBS, output_dir=tmp)
            assert os.path.getsize(result["excel_path"]) > 1024

    def test_build_report_pack_json_nontrivial_size(self):
        with tempfile.TemporaryDirectory() as tmp:
            result = build_report_pack(all_jobs=SAMPLE_JOBS, output_dir=tmp)
            assert os.path.getsize(result["json_path"]) > 256
