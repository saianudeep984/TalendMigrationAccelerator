"""
app/reports/excel_report.py
Generate an Excel migration report from analyzed job data.
Returns the file path of the written .xlsx file.
"""
from __future__ import annotations

import os
from typing import Any


def _v(val: Any) -> "str | int | float | bool | None":
    """Sanitize any value to an openpyxl-safe scalar."""
    if val is None or isinstance(val, (int, float, bool)):
        return val
    if isinstance(val, (list, dict, tuple, set)):
        return str(val)
    try:
        return str(val)
    except Exception:
        return ""


def export_excel(all_jobs: list[dict[str, Any]], output_dir: str = "output") -> str:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _fallback_csv(all_jobs, output_dir)

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "migration_report.xlsx")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _header_row(ws_sum, ["Metric", "Value"], row=1)
    total_jobs = len(all_jobs)
    total_components = sum(len(j.get("job_data", {}).get("components", [])) for j in all_jobs)
    total_high_risk = sum(
        1 for j in all_jobs for r in j.get("enterprise_risk_report", [])
        if r.get("risk") in ("HIGH", "CRITICAL")
    )
    auto_migratable = sum(1 for j in all_jobs if j.get("transformations"))
    auto_pct = round(100 * auto_migratable / total_jobs) if total_jobs else 0
    rows = [
        ("Total Jobs", total_jobs),
        ("Total Components", total_components),
        ("High/Critical Risk Findings", total_high_risk),
        ("Auto-Migratable Jobs", f"{auto_pct}%"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws_sum.cell(row=i, column=1, value=_v(k))
        ws_sum.cell(row=i, column=2, value=_v(str(v)))
    _autofit(ws_sum)

    # ── Sheet 2: Jobs ────────────────────────────────────────────────────────
    ws_jobs = wb.create_sheet("Jobs")
    job_cols = ["Job Name", "Version", "Components", "Complexity", "Cloud RAG",
                "High Risk", "Auto-Migratable", "Est. Hours"]
    _header_row(ws_jobs, job_cols, row=1)
    for i, job in enumerate(all_jobs, start=2):
        jd = job.get("job_data", {})
        cr = job.get("cloud_readiness", {})
        hi = sum(1 for r in job.get("enterprise_risk_report", []) if r.get("risk") in ("HIGH", "CRITICAL"))
        est = job.get("estimation", {})
        ws_jobs.cell(row=i, column=1, value=_v(jd.get("job_name", "")))
        ws_jobs.cell(row=i, column=2, value=_v(jd.get("job_version", "")))
        ws_jobs.cell(row=i, column=3, value=_v(len(jd.get("components", []))))
        ws_jobs.cell(row=i, column=4, value=_v(job.get("complexity", {}).get("level", "")))
        ws_jobs.cell(row=i, column=5, value=_v(cr.get("rag", "")))
        ws_jobs.cell(row=i, column=6, value=_v(hi))
        ws_jobs.cell(row=i, column=7, value="Yes" if job.get("transformations") else "No")
        ws_jobs.cell(row=i, column=8, value=_v(est.get("estimated_hours", "")))
    _autofit(ws_jobs)

    # ── Sheet 3: Risk Findings ───────────────────────────────────────────────
    ws_risk = wb.create_sheet("Risk Findings")
    risk_cols = ["Job", "Severity", "Component", "Type", "Detail"]
    _header_row(ws_risk, risk_cols, row=1)
    row_idx = 2
    for job in all_jobs:
        jname = job.get("job_data", {}).get("job_name", "")
        for r in job.get("enterprise_risk_report", []) + job.get("legacy_risk_report", []):
            ws_risk.cell(row=row_idx, column=1, value=_v(jname))
            ws_risk.cell(row=row_idx, column=2, value=_v(r.get("risk", r.get("severity", ""))))
            ws_risk.cell(row=row_idx, column=3, value=_v(r.get("component", "")))
            ws_risk.cell(row=row_idx, column=4, value=_v(r.get("type", r.get("category", ""))))
            ws_risk.cell(row=row_idx, column=5, value=_v(r.get("message", r.get("details", ""))))
            row_idx += 1
    _autofit(ws_risk)

    # ── Sheet 4: Components ──────────────────────────────────────────────────
    ws_comp = wb.create_sheet("Components")
    comp_cols = ["Job", "Component Name", "Component Type", "Version"]
    _header_row(ws_comp, comp_cols, row=1)
    row_idx = 2
    for job in all_jobs:
        jname = job.get("job_data", {}).get("job_name", "")
        for c in job.get("job_data", {}).get("components", []):
            ws_comp.cell(row=row_idx, column=1, value=_v(jname))
            ws_comp.cell(row=row_idx, column=2, value=_v(c.get("component_name", c.get("name", ""))))
            ws_comp.cell(row=row_idx, column=3, value=_v(c.get("component_type", c.get("type", ""))))
            ws_comp.cell(row=row_idx, column=4, value=_v(c.get("version", "")))
            row_idx += 1
    _autofit(ws_comp)

    # ── Sheet 5: Recommendations ─────────────────────────────────────────────
    ws_rec = wb.create_sheet("Recommendations")
    rec_cols = ["Job", "Recommendation", "Auto-Fix", "Priority"]
    _header_row(ws_rec, rec_cols, row=1)
    row_idx = 2
    for job in all_jobs:
        jname = job.get("job_data", {}).get("job_name", "")
        _recs = job.get("modernization_report", {})
        if isinstance(_recs, dict):
            _recs = _recs.get("recommendations", [])
        elif not isinstance(_recs, list):
            _recs = []
        for r in _recs:
            if isinstance(r, dict):
                _desc = _v(r.get("description", r.get("recommendation", r.get("message", str(r)))))
                _fix  = "Yes" if r.get("auto_fix") else "No"
                _pri  = _v(r.get("priority", ""))
            else:
                _desc = _v(r)
                _fix  = "No"
                _pri  = ""
            ws_rec.cell(row=row_idx, column=1, value=_v(jname))
            ws_rec.cell(row=row_idx, column=2, value=_desc)
            ws_rec.cell(row=row_idx, column=3, value=_fix)
            ws_rec.cell(row=row_idx, column=4, value=_pri)
            row_idx += 1
    _autofit(ws_rec)

    wb.save(out_path)
    return out_path


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_row(ws, cols: list[str], row: int = 1) -> None:
    try:
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.fill = fill
            cell.font = font
    except Exception:
        for col_idx, col_name in enumerate(cols, start=1):
            ws.cell(row=row, column=col_idx, value=col_name)


def _autofit(ws, min_width: int = 10, max_width: int = 60) -> None:
    try:
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(
                min_width, min(max_len + 2, max_width)
            )
    except Exception:
        pass


def _fallback_csv(all_jobs: list, output_dir: str) -> str:
    """Fallback: write a simple CSV if openpyxl is unavailable."""
    import csv
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "migration_report.csv")
    try:
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Job Name", "Version", "Components", "Complexity", "Cloud RAG", "High Risk"])
            for job in all_jobs:
                jd = job.get("job_data", {})
                cr = job.get("cloud_readiness", {})
                hi = sum(1 for r in job.get("enterprise_risk_report", []) if r.get("risk") in ("HIGH", "CRITICAL"))
                writer.writerow([
                    jd.get("job_name", ""),
                    jd.get("job_version", ""),
                    len(jd.get("components", [])),
                    job.get("complexity", {}).get("level", ""),
                    cr.get("rag", ""),
                    hi,
                ])
    except Exception:
        return ""
    return out_path
