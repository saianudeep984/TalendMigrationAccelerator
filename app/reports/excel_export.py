"""
Excel export for complete assessment reports (Repository Overview + Upgrade Path + all sections).
"""
from __future__ import annotations

from typing import Any, Dict, List, Sequence

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.tiap.documentation.export_utils import resolve_source_version as _resolve_source_version


def write_complete_assessment_excel(
    path: str,
    sections: Dict[str, str],
    all_jobs: Sequence[Dict[str, Any]] = (),
    repository_path: str | None = None,
) -> str:
    """Write a multi-sheet Excel workbook from report sections and job data."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    _write_repository_overview_sheet(wb, all_jobs, repository_path)
    _write_upgrade_path_sheet(wb, all_jobs, repository_path)

    # Remaining sections as sheets (truncated text)
    _SKIP = {"Repository Overview", "Upgrade Path"}
    for title, content in sections.items():
        if title in _SKIP:
            continue
        ws = wb.create_sheet(title=_safe_sheet_name(title))
        _write_section_sheet(ws, title, content)

    wb.save(path)
    return path


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names max 31 chars, no special chars."""
    invalid = r"\/*?:[]"
    clean = "".join(c for c in name if c not in invalid)
    return clean[:31]


def _header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1D4ED8")
    cell.alignment = Alignment(wrap_text=True)


def _write_repository_overview_sheet(wb: openpyxl.Workbook, all_jobs: Sequence[Dict[str, Any]], repository_path: str | None = None) -> None:
    ws = wb.create_sheet(title="Repository Overview")
    total_jobs = len(all_jobs)
    total_components = sum(len(j.get("job_data", {}).get("components", [])) for j in all_jobs)
    source_version = _resolve_source_version(all_jobs, repository_path)

    headers = ["KPI", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    rows = [
        ("Total Jobs", total_jobs),
        ("Total Components", total_components),
        ("Source Version", source_version),
        ("Target Version", "Talend 8"),
    ]
    for r, (kpi, val) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=kpi)
        ws.cell(row=r, column=2, value=val)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _write_upgrade_path_sheet(wb: openpyxl.Workbook, all_jobs: Sequence[Dict[str, Any]], repository_path: str | None = None) -> None:
    ws = wb.create_sheet(title="Upgrade Path")
    headers = ["Job Name", "Source Version", "Target Version", "Components", "Complexity"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)

    global_source = _resolve_source_version(all_jobs, repository_path)

    for r, job in enumerate(all_jobs, 2):
        jd = job.get("job_data", {})
        complexity = job.get("complexity", {})
        level = complexity.get("level") if isinstance(complexity, dict) else "UNKNOWN"
        ws.cell(row=r, column=1, value=jd.get("job_name", "Unknown"))
        ws.cell(row=r, column=2, value=jd.get("source_version", global_source))
        ws.cell(row=r, column=3, value="Talend 8")
        ws.cell(row=r, column=4, value=len(jd.get("components", [])))
        ws.cell(row=r, column=5, value=level or "UNKNOWN")

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 20


def _write_section_sheet(ws, title: str, content: str) -> None:
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    # Write truncated content (max 5000 chars per cell due to Excel limits)
    text = str(content or "")
    chunk_size = 5000
    row = 3
    for i in range(0, min(len(text), 30000), chunk_size):
        ws.cell(row=row, column=1, value=text[i: i + chunk_size])
        row += 1
    ws.column_dimensions["A"].width = 100
