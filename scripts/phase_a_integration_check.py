"""
Phase A Integration Validation
================================
Exercises, end-to-end, the full Phase A surface area:
  - RepositoryTypeDetector
  - EnterpriseFeatureDetector (EnterpriseFeatures)
  - UnsupportedComponentsAnalyzer
  - VersionCompatibilityEngine
  - UpgradePathAnalyzer
  - RepositoryOverviewCard (render smoke-test, headless)
  - Exports: DOCX / HTML / PDF / Excel / JSON (RepositoryOverview + UpgradePath data)

Run:
    python3 scripts/phase_a_integration_check.py
Exit code 0 == all checks passed.
"""
import json
import os
import sys
import tempfile
import traceback

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

RESULTS = []


def check(name):
    def decorator(fn):
        def wrapper():
            try:
                fn()
                RESULTS.append((name, True, ""))
                print(f"[PASS] {name}")
            except Exception as e:
                RESULTS.append((name, False, f"{type(e).__name__}: {e}"))
                print(f"[FAIL] {name} -> {type(e).__name__}: {e}")
                traceback.print_exc()
        return wrapper
    return decorator


SAMPLE_JOBS = [
    {
        "job_data": {
            "job_name": "Job_Alpha",
            "source_version": "Talend 6",
            "components": [
                {"component_type": "tMap"},
                {"component_type": "tFTPGet"},
                {"component_type": "tJavaFlex"},
            ],
        }
    },
    {
        "job_data": {
            "job_name": "Job_Beta",
            "source_version": "Talend 7",
            "components": [
                {"component_type": "tMap"},
                {"component_type": "tLogRow"},
            ],
        }
    },
]


@check("RepositoryTypeDetector.detect_from_path")
def t1():
    from app.repository.repository_type_detector import RepositoryTypeDetector
    d = RepositoryTypeDetector()
    result = d.detect_from_path("/tmp/Talend_6_Repository")
    assert "type" in result


@check("RepositoryTypeDetector.extract_source_version_from_path")
def t2():
    from app.repository.repository_type_detector import RepositoryTypeDetector
    d = RepositoryTypeDetector()
    v = d.extract_source_version_from_path("/tmp/Talend_6_Repository")
    assert v is not None


@check("EnterpriseFeatureDetector.detect_from_jobs")
def t3():
    from app.repository.enterprise_feature_detector import EnterpriseFeatureDetector
    info = EnterpriseFeatureDetector().detect_from_jobs(SAMPLE_JOBS)
    assert isinstance(info, dict)


@check("UnsupportedComponentsAnalyzer basic analysis")
def t4():
    from app.analyzers.unsupported_components_analyzer import UnsupportedComponentsAnalyzer
    analyzer = UnsupportedComponentsAnalyzer()
    assert hasattr(analyzer, "analyze"), "analyze() method not found on UnsupportedComponentsAnalyzer"
    result = analyzer.analyze(SAMPLE_JOBS)
    assert isinstance(result, dict)


@check("VersionCompatibilityEngine.get_supported_targets")
def t5():
    from app.analyzers.version_compatibility_engine import VersionCompatibilityEngine
    targets = VersionCompatibilityEngine().get_supported_targets("Talend 6")
    assert isinstance(targets, list)


@check("UpgradePathAnalyzer.build_hops / analyze_job / analyze_path")
def t6():
    from app.analyzers.version_upgrade_analyzer import UpgradePathAnalyzer
    a = UpgradePathAnalyzer()
    assert a.build_hops("Talend 6", "Talend 8") == ["6_to_7", "7_to_8"]
    result = a.analyze_job(SAMPLE_JOBS[0]["job_data"], "Talend 6", "Talend 8")
    assert "componentFindings" in result
    path = a.analyze_path("Talend 6", "Talend 8")
    assert "hops" in path


@check("RepositoryOverview model build + RepositoryOverviewCard render (headless)")
def t7():
    from app.analyzers.models import RepositoryOverview
    from app.ui.design_system_v2 import RepositoryOverviewCard
    overview = RepositoryOverview(
        total_jobs=2, total_components=5, total_routines=1, total_joblets=0,
        complexity_score=42, migration_readiness_score=65, cloud_readiness_score=70,
        testing_readiness_score=50, repository_type="Standard", source_version="Talend 6",
        enterprise_features=["tMDM", "tESB"], target_versions=["Talend 8"],
        migration_risk="MEDIUM",
        upgrade_path_summary="3 component change(s) required to upgrade from Talend 6 to Talend 8.",
    )
    # Headless render smoke test - streamlit calls are no-ops without a runtime
    # but must not raise.
    RepositoryOverviewCard(overview)


@check("build_report_pack_sections includes Repository Overview + Upgrade Path")
def t8():
    from app.tiap.documentation.report_pack_generator import build_report_pack_sections
    sections = build_report_pack_sections(SAMPLE_JOBS)
    assert "Repository Overview" in sections
    assert "Upgrade Path" in sections
    assert "Total Jobs" in sections["Repository Overview"]
    assert "Job_Alpha" in sections["Upgrade Path"]


@check("HTML export contains RepositoryOverview + UpgradePath data")
def t9():
    from app.tiap.documentation.report_pack_generator import (
        build_report_pack_sections,
        write_complete_assessment_html,
    )
    sections = build_report_pack_sections(SAMPLE_JOBS)
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
        path = f.name
    try:
        write_complete_assessment_html(path, sections)
        with open(path, encoding="utf-8") as fh:
            html = fh.read()
        assert "Repository Overview" in html
        assert "Upgrade Path" in html
        assert "<table" in html
        assert "Job_Alpha" in html
    finally:
        os.path.exists(path) and os.unlink(path)


@check("PDF export contains RepositoryOverview + UpgradePath text")
def t10():
    from app.tiap.documentation.report_pack_generator import (
        build_report_pack_sections,
        write_complete_assessment_pdf,
    )
    sections = build_report_pack_sections(SAMPLE_JOBS)
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        path = f.name
    write_complete_assessment_pdf(path, sections)
    assert os.path.exists(path)
    assert os.path.getsize(path) > 1024
    with open(path, "rb") as fh:
        assert fh.read(8).startswith(b"%PDF-")
    from pypdf import PdfReader
    reader = PdfReader(path)
    text = "".join(page.extract_text() or "" for page in reader.pages)
    assert "Repository Overview" in text
    assert "Upgrade Path" in text
    os.path.exists(path) and os.unlink(path)


@check("Excel export contains RepositoryOverview + UpgradePath sheets")
def t11():
    from app.tiap.documentation.report_pack_generator import build_report_pack_sections
    from app.reports.excel_export import write_complete_assessment_excel
    sections = build_report_pack_sections(SAMPLE_JOBS)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_complete_assessment_excel(path, sections, SAMPLE_JOBS)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 100
        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert "Repository Overview" in wb.sheetnames
        assert "Upgrade Path" in wb.sheetnames
    finally:
        os.path.exists(path) and os.unlink(path)


@check("JSON export contains RepositoryOverview + UpgradePath data")
def t12():
    from app.tiap.documentation.report_pack_generator import build_report_pack_sections
    from app.reports.json_export import write_complete_assessment_json
    sections = build_report_pack_sections(SAMPLE_JOBS)
    with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
        path = f.name
    try:
        write_complete_assessment_json(path, sections, SAMPLE_JOBS)
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        assert "repositoryOverview" in data
        assert "upgradePath" in data
        assert data["repositoryOverview"]["totalJobs"] == 2
        assert "jobs" in data["upgradePath"]
    finally:
        os.path.exists(path) and os.unlink(path)


@check("build_report_pack end-to-end (docx/html/pdf/excel/json)")
def t13():
    from app.tiap.documentation.report_pack_generator import build_report_pack
    with tempfile.TemporaryDirectory() as tmp:
        result = build_report_pack(all_jobs=SAMPLE_JOBS, output_dir=tmp)
        for key in ("docx_path", "html_path", "pdf_path", "excel_path", "json_path"):
            assert key in result, f"Missing {key} in build_report_pack result"
            assert os.path.exists(result[key]), f"{key} file not created"


def main():
    for fn in (t1, t2, t3, t4, t5, t6, t7, t8, t9, t10, t11, t12, t13):
        fn()
    print("\n--- Summary ---")
    passed = sum(1 for _, ok, _ in RESULTS if ok)
    total = len(RESULTS)
    for name, ok, err in RESULTS:
        print(f"{'PASS' if ok else 'FAIL'}: {name}" + (f" ({err})" if err else ""))
    print(f"\n{passed}/{total} checks passed")
    sys.exit(0 if passed == total else 1)


if __name__ == "__main__":
    main()
